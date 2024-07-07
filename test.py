import datetime
import openpyxl
import os
import smtplib
import ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl.styles import Alignment, PatternFill, Font

# Global variables for Excel workbook and worksheet
EXCEL_FILE = 'buyer_data.xlsx'
EXCEL_SHEET = 'Buyer Orders'

# User credentials and roles
USERS = {
    'jerald': {'password': 'l03e1t3', 'role': 'Working client'},
    'admin1': {'password': 'l03e1t3', 'role': 'Administrator'}
}

# Gmail configuration
GMAIL_USER = 'jcristorhen.10@gmail.com'
GMAIL_APP_PASSWORD = '@#$l03e1t3@#$JJ'  # Enter your Gmail app password here

# Function to authenticate user login
def login():
    while True:
        username = input("Enter username: ").strip()
        password = input("Enter password: ").strip()
        if username in USERS and USERS[username]['password'] == password:
            return username, USERS[username]['role']
        else:
            print("Invalid username or password. Please try again.")

# Function to add item to cart
def add_to_cart(item_number, inventory, cart, total):
    if item_number in inventory:
        item_name = inventory[item_number]['name']
        price = inventory[item_number]['price']
        if item_name in cart:
            cart[item_name]['quantity'] += 1
        else:
            cart[item_name] = {'quantity': 1, 'price': price}
        total[0] += price
        print(f"Added '{item_name}' to cart. Total: PHP {total[0]:,.2f}")
    else:
        print(f"Item with number {item_number} not found in inventory.")

# Function to remove item from cart
def remove_from_cart(item_index, cart, total):
    if item_index <= len(cart):
        item_name = list(cart.keys())[item_index - 1]
        if item_name in cart:
            cart[item_name]['quantity'] -= 1
            total[0] -= cart[item_name]['price']
            if cart[item_name]['quantity'] <= 0:
                del cart[item_name]
            print(f"Removed one '{item_name}' from cart. Total: PHP {total[0]:,.2f}")
        else:
            print(f"Item '{item_name}' is not in the cart.")
    else:
        print("Invalid item number. Please select a valid item number from the cart.")

# Function to clear all items from cart
def clear_cart(cart, total):
    cart.clear()
    total[0] = 0.0
    print("Cart cleared.")

# Function to generate receipt and store buyer data in Excel
def generate_receipt(cart, total, buyer_id):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print("\n===== Receipt =====")
    print(f"Date/Time: {timestamp}")
    print("-------------------")
    for item, details in cart.items():
        print(f"{item}: {details['quantity']}pcs. @ PHP {details['price']} = PHP {details['quantity'] * details['price']:.2f}")
    print("-------------------")
    print(f"Total: PHP {total[0]:,.2f}")
    print("===================")

    # Store buyer data in Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb[EXCEL_SHEET]

    row = (buyer_id, ', '.join([f"{item}: {details['quantity']}pcs." for item, details in cart.items()]), f"PHP {total[0]:,.2f}")
    sheet.append(row)

    # Formatting styles
    left_alignment = Alignment(horizontal='left', vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center')
    buyer_id_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue
    header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")  # Dark gray
    orders_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gray
    price_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red

    # Apply formatting
    for cell in sheet[f"A{sheet.max_row}:C{sheet.max_row}"]:
        for c in cell:
            if c.column == 1:
                c.alignment = left_alignment
                c.fill = buyer_id_fill
            elif c.column == 2:
                c.alignment = left_alignment
                c.fill = orders_fill
            elif c.column == 3:
                c.alignment = center_alignment
                c.fill = price_fill
            c.font = Font(bold=True)

    # Apply header formatting
    for col in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
        for cell in col:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")  # White font color for headers

    wb.save(EXCEL_FILE)
    wb.close()

# Function to load inventory from file and categorize items
def load_inventory(filename):
    inventory = {}
    category = None

    try:
        with open(filename, 'r') as file:
            for line in file:
                line = line.strip()
                if line:
                    if line.isupper():  # Assume it's a category name
                        category = line
                    else:
                        parts = line.split(',')
                        if len(parts) == 3:
                            item_number, name, price_str = parts
                            price = float(price_str.replace('$', '').strip())
                            if category:
                                inventory[item_number] = {'name': name, 'price': price, 'category': category}
                            else:
                                inventory[item_number] = {'name': name, 'price': price}
                        else:
                            print(f"Ignoring malformed line: {line}")
    except FileNotFoundError:
        print(f"Inventory file '{filename}' not found.")
    
    return inventory

# Function to initialize Excel file with headers if it doesn't exist
def initialize_excel():
    if not os.path.isfile(EXCEL_FILE):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = EXCEL_SHEET
        sheet.append(['Buyer ID', 'Items Ordered', 'Total'])
        wb.save(EXCEL_FILE)
        wb.close()

# Function to get the next available Buyer ID
def get_next_buyer_id():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb[EXCEL_SHEET]

    if sheet.max_row > 1:
        last_buyer_id = sheet.cell(row=sheet.max_row, column=1).value
        next_buyer_id = last_buyer_id + 1
    else:
        next_buyer_id = 1

    wb.close()
    return next_buyer_id

# Function to display the current cart contents
def display_cart(cart, total):
    print("\nCurrent Cart Contents:")
    if cart:
        for index, (item, details) in enumerate(cart.items(), start=1):
            print(f"{index}. {item}: {details['quantity']}pcs. @ PHP {details['price']} each")
        print(f"Total: PHP {total[0]:,.2f}")
    else:
        print("Cart is empty.")

# Function to send Gmail message (Administrator only)
def send_gmail_message():
    print("\n=== Gmail Message ===")
    recipient = input("Enter recipient's email address: ").strip()
    subject = input("Enter email subject: ").strip()
    message = input("Enter message: ").strip()

    try:
        # Create a secure SSL context
        context = ssl.create_default_context()

        # Connect to Gmail's SMTP server
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(GMAIL_USER, GMAIL_APP_PASSWORD)

            # Construct the email
            msg = MIMEMultipart()
            msg['From'] = GMAIL_USER
            msg['To'] = recipient
            msg['Subject'] = subject
            msg.attach(MIMEText(message, 'plain'))

            # Send email
            server.sendmail(GMAIL_USER, recipient, msg.as_string())

        print("Email sent successfully!")

    except Exception as e:
        print(f"Failed to send email. Error: {str(e)}")

# Main function to run the POS system
def main():
    inventory_file = 'inventory.txt'
    initialize_excel()  # Initialize Excel file with headers if it doesn't exist
    inventory = load_inventory(inventory_file)

    print("\nWelcome to the Enhanced POS System")

    while True:
        username, role = login()  # User login

        if role == 'Working client':
            print(f"\nWelcome {username} (Working client) to the Enhanced POS System")

            cart = {}
            total = [0.0]
            buyer_id = get_next_buyer_id()  # Initialize buyer ID

            while True:
                print("\nInventory by Categories:")
                categories = set(item['category'] for item in inventory.values() if 'category' in item)
                
                for category in categories:
                    print(f"\n{category}:")
                    for key, item in inventory.items():
                        if 'category' in item and item['category'] == category:
                            print(f"{key}. {item['name']} - PHP {item['price']:.2f}")

                display_cart(cart, total)  # Display the cart contents before prompting for item number

                item_number = input("\nEnter item number to add to cart (Press 'R' to remove an item, 'D' when done, 'C' to clear cart, 'exit' to exit): ").strip().upper()

                if item_number == 'D':
                    proceed = input("Proceed to tally and print receipt? (yes/no): ").strip().lower()
                    if proceed == 'yes':
                        if len(cart) > 0:
                            generate_receipt(cart, total, buyer_id)
                            input("Press Enter to clear the screen...")  # Pause to allow user to see the receipt
                            os.system('cls' if os.name == 'nt' else 'clear')  # Clear the screen
                            cart = {}
                            total[0] = 0.0
                            buyer_id += 1  # Increment buyer ID for the next buyer
                        else:
                            print("Cart is empty. Add items before tallying.")
                    elif proceed == 'no':
                        continue
                    else:
                        print("Invalid input. Please enter 'yes' or 'no'.")
                    
                elif item_number == 'EXIT':
                    confirm_exit = input("Do you want to proceed to exit? (yes/no): ").strip().lower()
                    if confirm_exit == 'yes':
                        break  # Exit the while loop and end the program
                    elif confirm_exit == 'no':
                        continue
                    else:
                        print("Invalid input. Please enter 'yes' or 'no'.")
                    
                elif item_number == 'R':
                    if cart:
                        try:
                            item_index_to_remove = int(input("Enter the number of the item to remove from cart: ").strip())
                            if 1 <= item_index_to_remove <= len(cart):
                                remove_from_cart(item_index_to_remove, cart, total)
                            else:
                                print("Invalid item number. Please select a valid item number from the cart.")
                        except ValueError:
                            print("Invalid input. Please enter a valid number.")
                    else:
                        print("Cart is empty. Nothing to remove.")
                
                elif item_number == 'C':
                    confirm_clear = input("Are you sure you want to clear all items from the cart? (yes/no): ").strip().lower()
                    if confirm_clear == 'yes':
                        clear_cart(cart, total)
                    elif confirm_clear == 'no':
                        continue
                    else:
                        print("Invalid input. Please enter 'yes' or 'no'.")
                
                elif item_number in inventory:
                    add_to_cart(item_number, inventory, cart, total)
                
                else:
                    print("Invalid item number. Please select a valid item number from the inventory.")

        elif role == 'Administrator':
            print(f"\nWelcome {username} (Administrator) to the Enhanced POS System")

            while True:
                print("\nAdministrator Menu:")
                print("1. Sales and Analytics Reports")
                print("2. Inventory Management")
                print("3. Multiple Payment Methods")
                print("4. Offline mode and synchronization")
                print("5. Promotions and Discounts")
                print("6. Multi-store support")
                print("7. Customizable Receipts")
                print("8. Feedback and Review Management")
                print("9. Message Admin using Gmail")
                print("10. Exit")

                admin_choice = input("\nEnter your choice: ").strip()

                if admin_choice == '1':
                    # Implement Sales and Analytics Reports functionality
                    print("Generating Sales and Analytics Reports...")

                elif admin_choice == '2':
                    # Implement Inventory Management functionality
                    print("Accessing Inventory Management...")

                elif admin_choice == '3':
                    # Implement Multiple Payment Methods functionality
                    print("Setting up Multiple Payment Methods...")

                elif admin_choice == '4':
                    # Implement Offline mode and synchronization functionality
                    print("Enabling Offline mode and synchronization...")

                elif admin_choice == '5':
                    # Implement Promotions and Discounts functionality
                    print("Managing Promotions and Discounts...")

                elif admin_choice == '6':
                    # Implement Multi-store support functionality
                    print("Configuring Multi-store support...")

                elif admin_choice == '7':
                    # Implement Customizable Receipts functionality
                    print("Customizing Receipts...")

                elif admin_choice == '8':
                    # Implement Feedback and Review Management functionality
                    print("Managing Feedback and Reviews...")

                elif admin_choice == '9':
                    # Implement Message Admin using Gmail functionality
                    send_gmail_message()

                elif admin_choice == '10':
                    break  # Exit from Administrator menu
                
                else:
                    print("Invalid choice. Please enter a valid option.")

        else:
            print("Unknown role. Access denied.")

if __name__ == "__main__":
    main()
