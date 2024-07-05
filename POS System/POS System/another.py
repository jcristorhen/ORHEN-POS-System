import datetime
import openpyxl
import os
from openpyxl.styles import Alignment, PatternFill, Font

# Global variable for Excel workbook and worksheet
EXCEL_FILE = 'buyer_data.xlsx'
EXCEL_SHEET = 'Buyer Orders'

# Function to add item to cart
def add_to_cart(item_number, inventory, cart, total):
    if item_number in inventory:
        item_name = inventory[item_number]['name']
        price = inventory[item_number]['price']
        if item_name in cart:
            cart[item_name]['quantity'] += 1
        else:
            cart[item_name] = {'quantity': 1, 'price': price}
        total[0] += price
        print(f"Added '{item_name}' to cart. Total: PHP {total[0]}")
    else:
        print(f"Item with number {item_number} not found in inventory.")

# Function to generate receipt and store buyer data in Excel
def generate_receipt(cart, total, buyer_id):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print("\n===== Receipt =====")
    print(f"Date/Time: {timestamp}")
    print("-------------------")
    for item, details in cart.items():
        print(f"{item}: {details['quantity']}pcs. @ PHP {details['price']} = PHP {details['quantity'] * details['price']}")
    print("-------------------")
    print(f"Total: PHP {total[0]}")
    print("===================")

    # Store buyer data in Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb[EXCEL_SHEET]

    # Formatting styles
    left_alignment = Alignment(horizontal='left', vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center')
    buyer_id_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue
    header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")  # Dark gray
    orders_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gray
    price_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red

    # Add row to sheet
    row = (buyer_id, ', '.join([f"{item}: {details['quantity']}pcs." for item, details in cart.items()]), f"PHP {total[0]}")
    sheet.append(row)

    # Apply formatting
    for cell in sheet[f"A{sheet.max_row}:C{sheet.max_row}"]:
        for c in cell:
            if c.column == 1:
                c.alignment = left_alignment
                c.fill = buyer_id_fill
            elif c.column == 2:
                c.alignment = left_alignment
                c.fill = orders_fill
            elif c.column == 3:
                c.alignment = center_alignment
                c.fill = price_fill
            c.font = Font(bold=True)

    # Apply header formatting
    for col in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
        for cell in col:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")  # White font color for headers

    wb.save(EXCEL_FILE)
    wb.close()

# Function to load inventory from file
def load_inventory(filename):
    inventory = {}
    try:
        with open(filename, 'r') as file:
            for line in file:
                if line.strip():  # Check if the line is not empty
                    parts = line.strip().split(',')
                    if len(parts) == 3:
                        item_number, name, price_str = parts
                        price = float(price_str.replace('$', '').strip())  # Remove $ and convert to float
                        inventory[item_number] = {'name': name, 'price': price}
                    else:
                        print(f"Ignoring malformed line: {line.strip()}")
    except FileNotFoundError:
        print(f"Inventory file '{filename}' not found.")
    return inventory

# Function to initialize Excel file with headers if it doesn't exist
def initialize_excel():
    if not os.path.isfile(EXCEL_FILE):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = EXCEL_SHEET
        sheet.append(['Buyer ID', 'Items Ordered', 'Total'])
        wb.save(EXCEL_FILE)
        wb.close()

# Function to get the next available Buyer ID
def get_next_buyer_id():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb[EXCEL_SHEET]

    # Check the last row to determine the next available Buyer ID
    if sheet.max_row > 1:
        last_buyer_id = sheet.cell(row=sheet.max_row, column=1).value
        next_buyer_id = last_buyer_id + 1
    else:
        next_buyer_id = 1

    wb.close()
    return next_buyer_id

# Main function to run the POS system
def main():
    inventory_file = 'inventory.txt'
    initialize_excel()  # Initialize Excel file with headers if it doesn't exist
    inventory = load_inventory(inventory_file)

    cart = {}
    total = [0.0]  # Using a list to pass total by reference

    print("\nWelcome to the Enhanced POS System")

    while True:
        print("\nInventory:")
        for key, item in inventory.items():
            print(f"{key}. {item['name']} - PHP {item['price']:.2f}")  # Display price with PHP sign and 2 decimal places

        item_number = input("\nEnter item number to add to cart (Press 'D' when done, 'exit' to exit): ").strip().upper()

        if item_number == 'D':
            proceed = input("Proceed to tally and print receipt? (yes/no): ").strip().lower()
            if proceed == 'yes':
                if len(cart) > 0:
                    buyer_id = get_next_buyer_id()  # Get the next available Buyer ID
                    generate_receipt(cart, total, buyer_id)
                    cart = {}
                    total[0] = 0.0
                else:
                    print("Cart is empty. Add items before tallying.")
            elif proceed == 'no':
                continue
            else:
                print("Invalid input. Please enter 'yes' or 'no'.")
        
        elif item_number == 'EXIT':
            confirm_exit = input("Do you want to proceed to exit? (yes/no): ").strip().lower()
            if confirm_exit == 'yes':
                break  # Exit the while loop and end the program
            elif confirm_exit == 'no':
                continue
            else:
                print("Invalid input. Please enter 'yes' or 'no'.")

        elif item_number in inventory:
            add_to_cart(item_number, inventory, cart, total)

        else:
            print("Invalid item number. Please select a valid item number from the inventory.")

    print("Exiting the program. Thank you for using our POS system.")

if __name__ == "__main__":
    main()
        