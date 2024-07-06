import datetime
import os
import openpyxl

# Global variables for Excel workbook and worksheet
EXCEL_FILE = 'buyer_data.xlsx'
EXCEL_SHEET = 'Buyer Orders'

# Function to add item to cart
def add_to_cart(item_number, inventory, cart, total):
    if item_number in inventory:
        item_name = inventory[item_number]['name']
        price = inventory[item_number]['price']
        if item_name in cart:
            cart[item_name]['quantity'] += 1
        else:
            cart[item_name] = {'quantity': 1, 'price': price}
        total[0] += price
        print(f"Added '{item_name}' to cart. Total: PHP {total[0]:,.2f}")
    else:
        print(f"Item with number {item_number} not found in inventory.")

# Function to generate receipt and store buyer data in Excel
def generate_receipt(cart, total, buyer_id):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print("\n===== Receipt =====")
    print(f"Date/Time: {timestamp}")
    print("-------------------")
    for item, details in cart.items():
        print(f"{item}: {details['quantity']}pcs. @ PHP {details['price']} = PHP {details['quantity'] * details['price']:.2f}")
    print("-------------------")
    print(f"Total: PHP {total[0]:,.2f}")
    print("===================")

    # Store buyer data in Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb[EXCEL_SHEET]

    row = (buyer_id, ', '.join([f"{item}: {details['quantity']}pcs." for item, details in cart.items()]), f"PHP {total[0]:,.2f}")
    sheet.append(row)

    wb.save(EXCEL_FILE)
    wb.close()

# Function to load inventory from file and categorize items
def load_inventory(filename):
    inventory = {}

    try:
        with open(filename, 'r') as file:
            for line in file:
                line = line.strip()
                if line:
                    parts = line.split(',')
                    if len(parts) == 3:
                        item_number, name, price_str = parts
                        price = float(price_str.replace('$', '').strip())
                        inventory[item_number] = {'name': name, 'price': price}
                    else:
                        print(f"Ignoring malformed line: {line}")
    except FileNotFoundError:
        print(f"Inventory file '{filename}' not found.")
    
    return inventory

# Function to initialize Excel file with headers if it doesn't exist
def initialize_excel():
    if not os.path.isfile(EXCEL_FILE):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = EXCEL_SHEET
        sheet.append(['Buyer ID', 'Items Ordered', 'Total'])
        wb.save(EXCEL_FILE)
        wb.close()

# Function to get the next available Buyer ID
def get_next_buyer_id():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb[EXCEL_SHEET]

    if sheet.max_row > 1:
        last_buyer_id = sheet.cell(row=sheet.max_row, column=1).value
        next_buyer_id = last_buyer_id + 1
    else:
        next_buyer_id = 1

    wb.close()
    return next_buyer_id

# Function to display the current cart contents
def display_cart(cart, total):
    print("\nCurrent Cart Contents:")
    if cart:
        for index, (item, details) in enumerate(cart.items(), start=1):
            print(f"{index}. {item}: {details['quantity']}pcs. @ PHP {details['price']} each")
        print(f"Total: PHP {total[0]:,.2f}")
    else:
        print("Cart is empty.")

# Main function to run the POS system
def main():
    inventory_file = 'inventory.txt'
    initialize_excel()  # Initialize Excel file with headers if it doesn't exist
    inventory = load_inventory(inventory_file)

    cart = {}
    total = [0.0]

    buyer_id = get_next_buyer_id()  # Initialize buyer ID

    print("\nWelcome to the Enhanced POS System")

    while True:
        print("\nInventory:")
        for item_number, item_info in inventory.items():
            print(f"{item_number}. {item_info['name']} - PHP {item_info['price']:.2f}")

        display_cart(cart, total)  # Display the cart contents before prompting for item number

        action = input("\nEnter item number to add to cart, 'R' to remove, 'D' when done, 'exit' to exit: ").strip().upper()

        if action == 'D':
            proceed = input("Proceed to tally and print receipt? (yes/no): ").strip().lower()
            if proceed == 'yes':
                if len(cart) > 0:
                    generate_receipt(cart, total, buyer_id)
                    input("Press Enter to clear the screen...")  # Pause to allow user to see the receipt
                    os.system('cls' if os.name == 'nt' else 'clear')  # Clear the screen
                    cart = {}
                    total[0] = 0.0
                    buyer_id += 1  # Increment buyer ID for the next buyer
                else:
                    print("Cart is empty. Add items before tallying.")
            elif proceed == 'no':
                continue
            else:
                print("Invalid input. Please enter 'yes' or 'no'.")
        
        elif action == 'EXIT':
            confirm_exit = input("Do you want to proceed to exit? (yes/no): ").strip().lower()
            if confirm_exit == 'yes':
                break  # Exit the while loop and end the program
            elif confirm_exit == 'no':
                continue
            else:
                print("Invalid input. Please enter 'yes' or 'no'.")

        elif action == 'R':
            if cart:
                try:
                    item_index_to_remove = int(input("Enter the number of the item to remove from cart: ").strip())
                    if 1 <= item_index_to_remove <= len(cart):
                        item_to_remove = list(cart.keys())[item_index_to_remove - 1]
                        total[0] -= cart[item_to_remove]['price']
                        if cart[item_to_remove]['quantity'] > 1:
                            cart[item_to_remove]['quantity'] -= 1
                        else:
                            del cart[item_to_remove]
                        print(f"Removed one '{item_to_remove}' from cart. Total: PHP {total[0]:,.2f}")
                    else:
                        print("Invalid item number. Please select a valid item number from the cart.")
                except ValueError:
                    print("Invalid input. Please enter a valid number.")
            else:
                print("Cart is empty. Nothing to remove.")
        
        elif action in inventory:
            add_to_cart(action, inventory, cart, total)

        else:
            print("Invalid input. Please select a valid item number from the inventory.")

if __name__ == "__main__":
    main()
