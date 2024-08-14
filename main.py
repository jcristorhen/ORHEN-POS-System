import datetime
import openpyxl
import os
import smtplib
import ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl.styles import Alignment, PatternFill, Font
import webbrowser

# Global variables for Excel workbook and worksheet
EXCEL_FILE = 'buyer_data.xlsx'
EXCEL_SHEET = 'Buyer Orders'
WORKING_CLIENTS_FILE = 'working_clients.xlsx'

# User credentials and roles
USERS = {
    'admin1': {'password': 'l03e1t3', 'role': 'Administrator'}
}

# Function to authenticate user login
def login(role=None):
    while True:
        username = input("Enter username: ").strip()
        password = input("Enter password: ").strip()

        # Check if user is in working clients
        try:
            wb = openpyxl.load_workbook(WORKING_CLIENTS_FILE)
            sheet = wb.active

            for row in sheet.iter_rows(values_only=True):
                if username == row[1] and password == row[2]:
                    return username, 'Working client'

            print("You are not registered. Please contact your admin. Thank you!")
        except FileNotFoundError:
            print("Working clients database not found.")
            return None, None

        # Check if user is in USERS dictionary
        if username in USERS and USERS[username]['password'] == password:
            if role and USERS[username]['role'] != role:
                print(f"Access denied. You are not authorized to access as {role}.")
                continue
            return username, USERS[username]['role']
        else:
            print("Invalid username or password. Please try again.")

# Function to add item to cart
def add_to_cart(item_number, inventory, cart, total):
    if item_number in inventory:
        item_name = inventory[item_number]['name']
        price = inventory[item_number]['price']
        if item_name in cart:
            cart[item_name]['quantity'] += 1
        else:
            cart[item_name] = {'quantity': 1, 'price': price}
        total[0] += price
        print(f"Added '{item_name}' to cart. Total: PHP {total[0]:,.2f}")
    else:
        print(f"Item with number {item_number} not found in inventory.")

# Function to remove item from cart
def remove_from_cart(item_index, cart, total):
    if item_index <= len(cart):
        item_name = list(cart.keys())[item_index - 1]
        if item_name in cart:
            cart[item_name]['quantity'] -= 1
            total[0] -= cart[item_name]['price']
            if cart[item_name]['quantity'] <= 0:
                del cart[item_name]
            print(f"Removed one '{item_name}' from cart. Total: PHP {total[0]:,.2f}")
        else:
            print(f"Item '{item_name}' is not in the cart.")
    else:
        print("Invalid item number. Please select a valid item number from the cart.")

# Function to clear all items from cart
def clear_cart(cart, total):
    cart.clear()
    total[0] = 0.0
    print("Cart cleared.")

# Function to generate receipt and store buyer data in Excel
def generate_receipt(cart, total, buyer_id):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print("\n===== Receipt =====")
    print(f"Date/Time: {timestamp}")
    print("-------------------")
    for item, details in cart.items():
        print(f"{item}: {details['quantity']}pcs. PHP {details['price']} = PHP {details['quantity'] * details['price']:.2f}")
    print("-------------------")
    print(f"Total: PHP {total[0]:,.2f}")
    print("===================")

    # Store buyer data in Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb[EXCEL_SHEET]

    row = (buyer_id, ', '.join([f"{item}: {details['quantity']}pcs." for item, details in cart.items()]), f"PHP {total[0]:,.2f}")
    sheet.append(row)

    # Formatting styles
    left_alignment = Alignment(horizontal='left', vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center')
    buyer_id_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue
    header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")  # Dark gray
    orders_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gray
    price_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red

    # Apply formatting
    for cell in sheet[f"A{sheet.max_row}:C{sheet.max_row}"]:
        for c in cell:
            if c.column == 1:
                c.alignment = left_alignment
                c.fill = buyer_id_fill
            elif c.column == 2:
                c.alignment = left_alignment
                c.fill = orders_fill
            elif c.column == 3:
                c.alignment = center_alignment
                c.fill = price_fill
            c.font = Font(bold=True)

    # Apply header formatting
    for col in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
        for cell in col:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")  # White font color for headers

    wb.save(EXCEL_FILE)
    wb.close()

# Function to load inventory from file and categorize items
def load_inventory(filename):
    inventory = {}
    category = None

    try:
        with open(filename, 'r') as file:
            for line in file:
                line = line.strip()
                if line:
                    if line.isupper():  # Assume it's a category name
                        category = line
                    else:
                        parts = line.split(',')
                        if len(parts) == 3:
                            item_number, name, price_str = parts
                            price = float(price_str.replace('$', '').strip())
                            if category:
                                inventory[item_number] = {'name': name, 'price': price, 'category': category}
                            else:
                                inventory[item_number] = {'name': name, 'price': price}
                        else:
                            print(f"Ignoring malformed line: {line}")
    except FileNotFoundError:
        print(f"Inventory file '{filename}' not found.")
    
    return inventory

# Function to initialize Excel file with headers if it doesn't exist
def initialize_excel():
    if not os.path.isfile(EXCEL_FILE):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = EXCEL_SHEET
        sheet.append(['Buyer ID', 'Items Ordered', 'Total'])
        wb.save(EXCEL_FILE)
        wb.close()

# Function to get the next available Buyer ID
def get_next_buyer_id():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb[EXCEL_SHEET]

    if sheet.max_row > 1:
        last_buyer_id = sheet.cell(row=sheet.max_row, column=1).value
        next_buyer_id = last_buyer_id + 1
    else:
        next_buyer_id = 1

    wb.close()
    return next_buyer_id

# Function to display the current cart contents
def display_cart(cart, total):
    print("\nCurrent Cart Contents:")
    if cart:
        for index, (item, details) in enumerate(cart.items(), start=1):
            print(f"{index}. {item}: {details['quantity']}pcs. | PHP {details['price']} each")
        print(f"Total: PHP {total[0]:,.2f}")
    else:
        print("Cart is empty.")

# Function to add inventory items
def add_inventory(inventory):
    print("\n=== Add Inventory Item ===")
    item_number = input("Enter item number: ").strip()
    name = input("Enter item name: ").strip()
    price = float(input("Enter item price: ").strip())
    category = input("Enter item category: ").strip().upper()
    inventory[item_number] = {'name': name, 'price': price, 'category': category}
    print(f"Item '{name}' added to inventory.")

# Function to remove inventory items
def remove_inventory(inventory):
    print("\n=== Remove Inventory Item ===")
    item_number = input("Enter item number to remove: ").strip()
    if item_number in inventory:
        del inventory[item_number]
        print(f"Item with number {item_number} removed from inventory.")
    else:
        print(f"Item with number {item_number} not found in inventory.")

# Function to register a new working client
def register_working_client():
    print("\n=== Register New Working Client ===")
    username = input("Enter new username: ").strip()
    password = input("Enter new password: ").strip()

    wb = openpyxl.load_workbook(WORKING_CLIENTS_FILE)
    sheet = wb.active

    # Find the last row number with data
    last_row = sheet.max_row

    # Append the new data below the last row
    sheet.append((last_row + 1, username, password))

    wb.save(WORKING_CLIENTS_FILE)
    wb.close()

    print(f"Working client '{username}' registered successfully.")


def remove_working_client():
    print("\n=== Remove Working Client Account ===")
    username_to_remove = input("Enter username to remove: ").strip()

    wb = openpyxl.load_workbook(WORKING_CLIENTS_FILE)
    sheet = wb.active

    found = False
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[1].value == username_to_remove:
            sheet.delete_rows(row[0].row)
            found = True
            break

    if found:
        print(f"Working client '{username_to_remove}' account removed successfully.")
    else:
        print(f"Working client '{username_to_remove}' not found.")

    wb.save(WORKING_CLIENTS_FILE)
    wb.close()


# Function to view sales records
def view_sales_records():
    print("\n=== View Sales Records ===")

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb[EXCEL_SHEET]

    for row in sheet.iter_rows(values_only=True):
        print(f"Buyer ID: {row[0]} | Items Ordered: {row[1]} | Total: {row[2]}")

    wb.close()

# Function to clear the console screen
def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')

# Main POS system function
def pos_system(username, role):
    inventory_file = 'inventory.txt'
    initialize_excel()  # Initialize Excel file with headers if it doesn't exist
    inventory = load_inventory(inventory_file)

    if role == 'Working client':
        print(f"\nWelcome {username}! Login date: {datetime}")

        cart = {}
        total = [0.0]
        buyer_id = get_next_buyer_id()  # Initialize buyer ID

        while True:
            print("\nInventory by Categories:")
            categories = set(item['category'] for item in inventory.values() if 'category' in item)
            
            for category in categories:
                print(f"\n{category}:")
                for key, item in inventory.items():
                    if item.get('category') == category:
                        print(f"{key}. {item['name']} - PHP {item['price']:.2f}")

            display_cart(cart, total)  # Display the cart contents before prompting for item number

            item_number = input("\nEnter item number to add to cart: ").strip().upper()

            if item_number == 'D':
                proceed = input("Proceed to tally and print receipt? (yes/no): ").strip().lower()
                if proceed == 'yes':
                    if len(cart) > 0:
                        generate_receipt(cart, total, buyer_id)
                        input("Press Enter to clear the screen...")  # Pause to allow user to see the receipt
                        clear_screen()  # Clear the screen
                        cart = {}
                        total[0] = 0.0
                        buyer_id += 1  # Increment buyer ID for the next buyer
                    else:
                        print("Cart is empty. Add items before tallying.")
                elif proceed == 'no':
                    continue
                else:
                    print("Invalid input. Please enter 'yes' or 'no'.")
                
            elif item_number == 'EXIT':
                confirm_exit = input("Do you want to proceed to exit? (yes/no): ").strip().lower()
                if confirm_exit == 'yes':
                    return  # Exit the program by returning from the main function
                elif confirm_exit == 'no':
                    continue
                else:
                    print("Invalid input. Please enter 'yes' or 'no'.")
                
            elif item_number == 'R':
                if cart:
                    try:
                        item_index_to_remove = int(input("Enter the number of the item to remove from cart: ").strip())
                        if 1 <= item_index_to_remove <= len(cart):
                            remove_from_cart(item_index_to_remove, cart, total)
                        else:
                            print("Invalid item number. Please select a valid item number from the cart.")
                    except ValueError:
                        print("Invalid input. Please enter a valid number.")
                else:
                    print("Cart is empty. Nothing to remove.")
            
            elif item_number == 'C':
                confirm_clear = input("Are you sure you want to clear all items from the cart? (yes/no): ").strip().lower()
                if confirm_clear == 'yes':
                    clear_cart(cart, total)
                elif confirm_clear == 'no':
                    continue
                else:
                    print("Invalid input. Please enter 'yes' or 'no'.")
            
            elif item_number in inventory:
                add_to_cart(item_number, inventory, cart, total)
            
            else:
                print("Invalid item number. Please select a valid item number from the inventory.")

    elif role == 'Administrator':
        print("\nWELCOME BACK ADMIN!")

        while True:
            print("\nAdministrator Menu:")
            print("1. POS System")
            print("2. Inventory Management")
            print("3. Account Manager")
            print("4. Help")
            print("5. Exit")

            admin_choice = input("\nEnter your choice: ").strip()

            if admin_choice == '1':
                # Access the POS system for administrator
                pos_system('Administrator', 'Working client')  # Administrator can access POS as Working client

            elif admin_choice == '2':
                # Add or Remove Inventory
                while True:
                    print("\nInventory Management:")
                    print("1. Add Inventory")
                    print("2. Remove Inventory")
                    print("3. Back to Admin Menu")

                    inventory_choice = input("Enter your choice: ").strip()

                    if inventory_choice == '1':
                        add_inventory(inventory)
                    elif inventory_choice == '2':
                        remove_inventory(inventory)
                    elif inventory_choice == '3':
                        break
                    else:
                        print("Invalid choice. Please enter a valid option.")

            elif admin_choice == '3':
                # Account Manager
                while True:
                    print("\nAccount Manager:")
                    print("1. Register Working Client")
                    print("2. Remove Working Client Account")
                    print("3. View Sales Records")
                    print("4. Back to Admin Menu")

                    account_choice = input("Enter your choice: ").strip()

                    if account_choice == '1':
                        register_working_client()
                    elif account_choice == '2':
                        remove_working_client()
                    elif account_choice == '3':
                        view_sales_records()
                    elif account_choice == '4':
                        break
                    else:
                        print("Invalid choice. Please enter a valid option.")

            elif admin_choice == '4':
                # Help - Open the specified Facebook page
                webbrowser.open('https://www.facebook.com/jcrist.orhen')

            elif admin_choice == '5':
                print("Exiting the system...")
                return  # Exit the program by returning from the main function

            else:
                print("Invalid choice. Please enter a valid option.")

    else:
        print("Unknown role. Access denied.")

# Main function to run the POS system
def main():
    print("=== Welcome to the ORHEN POS System ===")

    while True:
        print("\nMenu:")
        print("1. Login as Working Client")
        print("2. Login as Administrator")
        print("3. Exit")

        choice = input("\nEnter your choice: ").strip()

        if choice == '1':
            username, role = login('Working client')
            if role == 'Working client':
                pos_system(username, role)
        elif choice == '2':
            username, role = login('Administrator')
            if role == 'Administrator':
                pos_system(username, role)
        elif choice == '3':
            print("Exiting the system...")
            break
        else:
            print("Invalid choice. Please enter a valid option.")

if __name__ == "__main__":
    main()